from openpyxl import load_workbook  # 读写
#case_id 第一行 url 第三行  param 第5 method  6  Expected 7
from common import get_path
import json
class Cases:
    def __init__(self):      #创建手机容器
        self.case_id=None
        self.url=None
        self.param=None
        self.method=None
        self.Expected=None
        self.title=None
        self.module=None
class doexcel:
    def __init__(self,file_name):
        self.file_name=file_name
    def read_data(self,sheet_name):
        self.sheet_name = sheet_name
        wb = load_workbook(self.file_name)  # 打开表
        sheet = wb[self.sheet_name]  # 定位表单
        cases_1=[]
        for i in range(2,sheet.max_row+1):
            row_case=Cases()
            row_case.case_id=sheet.cell(row=i,column=1).value   ##case_id的值
            row_case.module = sheet.cell(row=i, column=2).value  ##module的值
            row_case.url=sheet.cell(row=i,column=3).value      #URL的值
            row_case.title= sheet.cell(row=i, column=4).value ##title的值
            row_case.param=sheet.cell(row=i,column=5).value    #param的值
            row_case.method=sheet.cell(row=i,column=6).value   #method的值
            row_case.Expected=sheet.cell(row=i,column=7).value  #Expected的值
            cases_1.append(row_case)
        return cases_1

    def write_back(self, sheet_name, row,result,testresult):  # 传入行，列，数值
        self.sheet_name = sheet_name
        wb = load_workbook(self.file_name)  # 打开表
        sheet = wb[self.sheet_name]  # 定位表单
        sheet.cell(row,8).value = result
        sheet.cell(row,9).value = testresult
        wb.save(self.file_name)
if __name__ == '__main__':
    t=doexcel(get_path.cases_path)

    u=t.read_data('info')
    for i in u:
     print(i.module)
     t=json.loads(i.module)
     print(type(t))
